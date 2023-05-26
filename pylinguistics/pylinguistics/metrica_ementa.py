import Pylinguistics as pl
import glob
import natsort
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


class MetricaEmenta:

    sentence_path = ""

    def __init__(self, path):
        self.sentence_path = path

    def list_files_ordered(self, extension):
        files_br = []
        for file in glob.glob(self.sentence_path + extension):
            files_br.append(file)
        files_br = natsort.natsorted(files_br)
        return files_br

    def get_config_plan_xlsx(self, wb):

        ws = wb.active
        ws.title = 'metrics'

        ws['A1'] = 'EMENTA'
        ws['B1'] = 'QUANTIDADE DE SENTENÇAS'
        ws['C1'] = 'COMPRIMENTO MÉDIO DAS SENTENÇAS'
        ws['D1'] = 'NÚMERO MÉDIO DE SÍLABAS POR PALAVRA'
        ws['E1'] = 'LEGIBILIDADE'

        return ws

    def generate_metrics_by_ementa(self, name_file, extension="/*.br"):

        files_br = self.list_files_ordered(extension)

        # Funciona apenas para *.xlsx
        wb = Workbook()
        ws = self.get_config_plan_xlsx(wb)

        for i, file in enumerate(files_br):
            row = i + 1
            number_ementa = file.replace(self.sentence_path + "/ementa_", "")
            number_ementa = number_ementa.replace(".br", "")
            number_ementa = number_ementa.replace(".simplificada.sp", "") # exp01
            number_ementa = number_ementa.replace(".sp", "") # demais exps

            with open(file, 'r', encoding='utf8') as file_text:
                ementa = file_text.read() # lê o arquivo inteiro
            json_metrics = self.calc_metrics(ementa)

            inc = row + 1
            a = get_column_letter(1)
            ws[a + str(inc)] = number_ementa
            b = get_column_letter(2)
            ws[b + str(inc)] = json_metrics['sentence_count']
            c = get_column_letter(3)
            ws[c + str(inc)] = self.format_value(json_metrics['mean_sentence_length'])
            d = get_column_letter(4)
            ws[d + str(inc)] = self.format_value(json_metrics['avg_syllables_per_word'])
            e = get_column_letter(5)
            ws[e + str(inc)] = self.format_value(json_metrics['redability'])

            print(f'### MÉTRICA EMENTA {number_ementa} - ({extension}) CONCLUÍDA ###')

        wb.save(self.sentence_path + "/" + name_file)

    def calc_metrics(self, ementa):
        objpl = pl.text(ementa)
        objpl.setLanguage("pt-br")
        return objpl.getFeatures()

    def format_value(self, value):
        return "{:.2f}".format(value)
